import mcp.types as types
import json
import sys
import os
import re
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Union

# Excel and CSV libraries
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd

# Add the parent directory to the path
sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..'))
from mcp_server.pipeline.expressions import substitute_expressions, parse_substituted_string, build_default_context

async def tool_info() -> dict:
    """Returns information about the lng_xls_batch tool."""
    return {
        "name": "lng_xls_batch",
        "description": """Advanced Excel/CSV Batch Operations Tool

**Core Features:**
• **Multi-format Support** - Excel (.xlsx, .xls) and CSV files
• **Template-based Operations** - Copy from source to template, save as result file
• **Smart Copy Operations** - Values, formulas, formatting with flexible combinations
• **Expression System** - Dynamic ranges and values using {! expression !} syntax  
• **Batch Processing** - Multiple operations with automatic file saving
• **Flexible Insert Modes** - Replace or insert with row/column expansion
• **CSV Integration** - CSV files treated as single-sheet Excel internally

**Operation Modes:**
• `single` - Single copy operation (auto-detected)
• `batch` - Multiple operations with workspace management

**Copy Types:**
• `values` - Copy cell values only
• `formulas` - Copy formulas only
• `formatting` - Copy cell formatting only
• Combinations: `["values", "formulas"]`, `["values", "formatting"]`, etc.

**Workspace Management:**
Define file aliases for easy reference:
```json
{
  "workspace": {
    "source": "input.xlsx",
    "target": "output.csv",
    "temp": "calculations.xlsx"
  }
}
```

**Insert Modes:**
• `"replace"` - Replace existing cells
• `["rows"]` - Insert new rows, shifting existing down
• `["columns"]` - Insert new columns, shifting existing right  
• `["rows", "columns"]` - Insert both rows and columns

**Range Notation:**
• Excel files: `[fileId]SheetName!A1:C10`
• CSV files: `[fileId]A1:C10` (sheet name optional)
• Dynamic ranges: `[source]Sheet1!A1:A{! row_count !}`
• Template operations: copy from `[source]` to `[template]`, save result as `[result]`

**Example Usage:**

**Template-based Copy:**
```json
{
  "workspace": {
    "source": "data.xlsx",
    "template": "template.xlsx", 
    "result": "output.xlsx"
  },
  "operations": [
    {
      "from": "[source]Sheet1!A1:C10",
      "to": "[template]Summary!A1:C10",
      "save_as": "result",
      "copy": ["values", "formulas"],
      "insert": "replace"
    }
  ]
}
```

**Batch with Expressions:**
```json
{
  "workspace": {
    "input": "raw_data.csv", 
    "template": "report_template.xlsx",
    "output": "processed_report.xlsx"
  },
  "defaults": {
    "copy": ["values"],
    "insert": "replace"
  },
  "operations": [
    {
      "from": "{! env.REPORT_TITLE !}",
      "to": "[template]Report!A1",
      "save_as": "output"
    },
    {
      "from": "=SUM(B:B)",
      "to": "[template]Report!B{! row_count + 2 !}",
      "save_as": "output",
      "copy": ["formulas"]
    }
  ]
}
```

**Processing Flow:**
1. Load workspace files (create if needed)
2. Process expressions in operations
3. Copy from source to template (template remains unchanged)
4. Save modified template as result file specified in save_as
5. Handle CSV ↔ Excel conversion automatically""",
        
        "schema": {
            "type": "object",
            "properties": {
                "workspace": {
                    "type": "object",
                    "description": "File workspace with aliases for easy reference",
                    "additionalProperties": {"type": "string"}
                },
                "defaults": {
                    "type": "object", 
                    "description": "Default settings for operations",
                    "properties": {
                        "copy": {
                            "type": "array",
                            "items": {"type": "string", "enum": ["values", "formulas", "formatting"]},
                            "default": ["values", "formulas"]
                        },
                        "insert": {
                            "oneOf": [
                                {"type": "string", "enum": ["replace"]},
                                {"type": "array", "items": {"type": "string", "enum": ["rows", "columns"]}}
                            ],
                            "default": "replace"
                        }
                    }
                },
                "operations": {
                    "type": "array",
                    "description": "List of copy operations to perform",
                    "items": {
                        "type": "object",
                        "properties": {
                            "from": {
                                "type": "string",
                                "description": "Source range, formula, or expression"
                            },
                            "to": {
                                "type": "string", 
                                "description": "Target template range with file and sheet reference"
                            },
                            "save_as": {
                                "type": "string",
                                "description": "File ID from workspace where to save the result (required)"
                            },
                            "copy": {
                                "type": "array",
                                "items": {"type": "string", "enum": ["values", "formulas", "formatting"]},
                                "description": "What to copy (overrides defaults)"
                            },
                            "insert": {
                                "oneOf": [
                                    {"type": "string", "enum": ["replace"]},
                                    {"type": "array", "items": {"type": "string", "enum": ["rows", "columns"]}}
                                ],
                                "description": "Insert mode (overrides defaults)"
                            }
                        },
                        "required": ["from", "to", "save_as"]
                    }
                },
                "debug": {
                    "type": "boolean",
                    "default": False,
                    "description": "Enable debug logging"
                },
                "analyze": {
                    "type": "boolean",
                    "default": False,
                    "description": "Analyze workspace files and return statistics without performing operations"
                }
            },
            "required": ["workspace"]
        }
    }

async def run_tool(name: str, parameters: dict) -> list[types.Content]:
    """Run tool function for MCP server."""
    try:
        result = await main(parameters)
        return [types.TextContent(type="text", text=json.dumps(result, indent=2))]
    except Exception as e:
        error_result = {"success": False, "error": str(e)}
        return [types.TextContent(type="text", text=json.dumps(error_result, indent=2))]

async def main(params: dict) -> dict:
    """Main function for lng_xls_batch tool."""
    try:
        # Set up logging
        logger = logging.getLogger(__name__)
        if params.get("debug", False):
            logger.setLevel(logging.DEBUG)
            
        # Build context for expressions
        context = build_default_context()
        
        # Extract parameters
        workspace = params.get("workspace", {})
        defaults = params.get("defaults", {
            "copy": ["values", "formulas"],
            "insert": "replace"
        })
        operations = params.get("operations", [])
        analyze_mode = params.get("analyze", False)  # New analyze mode
        
        # If analyze mode and no operations, return file statistics
        if analyze_mode and not operations:
            return analyze_workspace_files(workspace, context)
        
        if not analyze_mode and not operations:
            return {
                "success": False,
                "error": "No operations specified"
            }
            
        # Validate workspace
        if not workspace:
            return {
                "success": False, 
                "error": "Workspace configuration is required"
            }
            
        # Initialize file handlers
        file_handlers = {}
        
        # Process operations
        results = []
        save_as_files = {}  # Track files that need to be saved with save_as
        
        for i, operation in enumerate(operations):
            try:
                logger.info(f"Processing operation {i+1}/{len(operations)}")
                
                # Validate required save_as parameter
                if "save_as" not in operation:
                    raise ValueError(f"Operation {i+1} missing required 'save_as' parameter")
                
                save_as_id = operation["save_as"]
                if save_as_id not in workspace:
                    raise ValueError(f"Operation {i+1}: save_as file ID '{save_as_id}' not found in workspace")
                
                # Process expressions in operation
                processed_op = process_expressions(operation, context)
                
                # Execute operation
                result = await execute_operation(
                    processed_op, workspace, defaults, file_handlers, logger, context, save_as_files
                )
                
                results.append({
                    "operation": i + 1,
                    "success": True,
                    "result": result,
                    "save_as": save_as_id
                })
                
            except Exception as e:
                error_msg = f"Operation {i+1} failed: {str(e)}"
                logger.error(error_msg)
                return {
                    "success": False,
                    "error": error_msg,
                    "operation": i + 1,
                    "completed_operations": results
                }
        
        # Final save of all save_as files
        save_results = save_save_as_files(save_as_files, workspace, logger)
        
        return {
            "success": True,
            "operations_completed": len(results),
            "results": results,
            "files_saved": save_results
        }
        
    except Exception as e:
        logger.error(f"Tool execution failed: {str(e)}")
        return {
            "success": False,
            "error": f"Tool execution failed: {str(e)}"
        }

def process_expressions(operation: dict, context: dict) -> dict:
    """Process expressions in operation parameters."""
    processed = {}
    
    for key, value in operation.items():
        if isinstance(value, str):
            # Process expressions in string values
            processed[key] = substitute_expressions(value, context, "python")
        else:
            processed[key] = value
            
    return processed

async def execute_operation(operation: dict, workspace: dict, defaults: dict, 
                          file_handlers: dict, logger: logging.Logger, context: dict, save_as_files: dict) -> dict:
    """Execute a single copy operation."""
    
    from_source = operation["from"]
    to_target = operation["to"] 
    save_as_id = operation["save_as"]
    copy_types = operation.get("copy", defaults.get("copy", ["values", "formulas"]))
    insert_mode = operation.get("insert", defaults.get("insert", "replace"))
    
    logger.debug(f"Executing operation: {from_source} => {to_target} (save as {save_as_id})")
    
    # Parse source and target
    source_info = parse_range(from_source, workspace, context, logger)
    target_info = parse_range(to_target, workspace, context, logger)
    
    # Validate target must be a range (file reference)
    if target_info["type"] != "range":
        raise ValueError(f"Target must be a file range, got: {to_target}")
    
    # Load source file
    if source_info["type"] == "range":
        source_wb = load_workbook(source_info["file"], file_handlers)
        source_ws = get_worksheet(source_wb, source_info.get("sheet"))
    else:
        source_wb = None
        source_ws = None
    
    # Load template file (to_target file)
    template_wb = load_workbook(target_info["file"], file_handlers)
    template_ws = get_worksheet(template_wb, target_info.get("sheet"), create=True)
    
    # Create or get save_as workbook
    save_as_file_path = workspace[save_as_id]
    if save_as_id not in save_as_files:
        # First operation for this save_as
        if target_info["file"] == save_as_file_path:
            # save_as is same as template file - work directly with template
            save_as_wb = template_wb
        else:
            # save_as is different file - copy template file first
            import shutil
            shutil.copy2(target_info["file"], save_as_file_path)
            save_as_wb = load_workbook(save_as_file_path, {})
            
        save_as_files[save_as_id] = {
            'workbook': save_as_wb,
            'file_path': save_as_file_path
        }
    else:
        # Use existing save_as workbook
        save_as_wb = save_as_files[save_as_id]['workbook']
    
    # Get worksheet from save_as workbook
    save_as_ws = get_worksheet(save_as_wb, target_info.get("sheet"), create=True)
    
    # Execute copy operation based on source type - copy to save_as worksheet
    if source_info["type"] == "range":
        # Copy range
        result = copy_range(
            source_ws, source_info["range"],
            save_as_ws, target_info["range"], 
            copy_types, insert_mode, logger
        )
    elif source_info["type"] == "formula":
        # Copy formula
        result = copy_formula(
            source_info["formula"],
            save_as_ws, target_info["range"],
            copy_types, insert_mode, logger
        )
    elif source_info["type"] == "value":
        # Copy value
        result = copy_value(
            source_info["value"],
            save_as_ws, target_info["range"],
            copy_types, insert_mode, logger
        )
    else:
        raise ValueError(f"Unknown source type: {source_info['type']}")
    
    # Note: template file remains unchanged, only save_as file is modified
    
    return result

def parse_range(range_str: str, workspace: dict, context: dict = None, logger = None) -> dict:
    """Parse range string into components."""
    # Handle different source types
    if range_str.startswith("="):
        # Formula
        return {
            "type": "formula",
            "formula": range_str
        }
    elif range_str.startswith("{!") and range_str.endswith("!}"):
        # Expression value  
        return {
            "type": "value",
            "value": range_str
        }
    elif "[" in range_str and "]" in range_str:
        # File reference with range
        # Pattern: [fileId]Sheet!Range or [fileId]Range
        logger.debug(f"Attempting to parse range string: '{range_str}'")
        match = re.match(r'\[([a-zA-Z0-9_-]+)\](.+)', range_str)
        logger.debug(f"Regex match result: {match}")
        if match:
            logger.debug(f"Match groups: file_id='{match.group(1)}', rest='{match.group(2)}'")
        if not match:
            available_files = list(workspace.keys()) if workspace else []
            raise ValueError(f"Invalid range format: '{range_str}'. Expected format: '[fileId]Range' or '[fileId]Sheet!Range'. Available file IDs: {available_files}. Debug: range_str='{range_str}', len={len(range_str)}")
            
        file_id = match.group(1)
        rest = match.group(2)
        
        if file_id not in workspace:
            available_files = list(workspace.keys())
            raise ValueError(f"File ID '{file_id}' not found in workspace. Available file IDs: {available_files}. Range string: '{range_str}'")
            
        file_path = workspace[file_id]
        
        # Check if it has sheet name
        if "!" in rest:
            sheet_name, range_part = rest.split("!", 1)
            
            # Process expressions in range_part if context is provided
            if context and ("{!" in range_part or "[!" in range_part):
                range_part = substitute_expressions(range_part, context, "python")
            
            return {
                "type": "range",
                "file": file_path,
                "sheet": sheet_name,
                "range": range_part
            }
        else:
            # No sheet name (CSV case)
            range_part = rest
            
            # Process expressions in range_part if context is provided
            if context and ("{!" in range_part or "[!" in range_part):
                range_part = substitute_expressions(range_part, context, "python")
            
            return {
                "type": "range", 
                "file": file_path,
                "sheet": None,
                "range": range_part
            }
    else:
        # Simple value (fallback)
        return {
            "type": "value",
            "value": range_str
        }

def load_workbook(file_path: str, file_handlers: dict):
    """Load workbook from file or cache."""
    if file_path in file_handlers:
        return file_handlers[file_path]
    
    path = Path(file_path)
    
    if path.suffix.lower() == '.csv':
        # Load CSV as Excel
        wb = load_csv_as_excel(file_path)
    elif path.exists():
        # Load existing Excel file
        wb = openpyxl.load_workbook(file_path)
    else:
        # Create new Excel file
        wb = openpyxl.Workbook()
        
    file_handlers[file_path] = wb
    return wb

def load_csv_as_excel(csv_path: str):
    """Load CSV file as Excel workbook."""
    # Read CSV with pandas
    df = pd.read_csv(csv_path)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Write data to Excel
    for r_idx, row in enumerate(df.itertuples(index=False), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + 1, column=c_idx, value=value)  # +1 for header
    
    # Write headers
    for c_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=c_idx, value=col_name)
    
    return wb

def get_worksheet(wb, sheet_name: Optional[str], create: bool = False):
    """Get worksheet from workbook."""
    if sheet_name is None:
        # Return active sheet (CSV case)
        return wb.active
    
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    elif create:
        # Create new sheet
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        else:
            return wb.create_sheet(sheet_name)
    else:
        # For new workbooks, first sheet might not be named "Sheet1"
        if len(wb.sheetnames) == 1 and sheet_name == "Sheet1":
            # Rename the active sheet to the requested name
            wb.active.title = sheet_name
            return wb.active
        raise ValueError(f"Sheet '{sheet_name}' not found")

def calculate_target_range(target_range: str, source_cells) -> str:
    """Calculate proper target range based on source dimensions."""
    from openpyxl.utils import coordinate_to_tuple, get_column_letter
    
    # If target_range already contains ":" it's a full range - leave as is
    if ":" in target_range:
        return target_range
    
    # Parse starting position
    start_row, start_col = coordinate_to_tuple(target_range)
    
    # Calculate dimensions needed
    if hasattr(source_cells, 'coordinate'):
        # Single cell - target is also single cell
        return target_range
    elif isinstance(source_cells, tuple) and len(source_cells) > 0 and isinstance(source_cells[0], tuple):
        # Multiple rows
        rows_needed = len(source_cells)
        cols_needed = len(source_cells[0])
    else:
        # Single row or multiple cells in one row
        rows_needed = 1
        cols_needed = len(source_cells) if isinstance(source_cells, tuple) else 1
    
    # Calculate end position
    end_row = start_row + rows_needed - 1
    end_col = start_col + cols_needed - 1
    
    # Build range string
    start_coord = target_range
    end_coord = f"{get_column_letter(end_col)}{end_row}"
    
    return f"{start_coord}:{end_coord}"

def copy_range(source_ws: Worksheet, source_range: str, 
              target_ws: Worksheet, target_range: str,
              copy_types: List[str], insert_mode, logger: logging.Logger) -> dict:
    """Copy range between worksheets."""
    # Parse ranges
    source_cells = source_ws[source_range]
    
    # Handle insert mode
    if insert_mode != "replace":
        logger.info(f"Handling insert mode: {insert_mode}")
        handle_insert_mode(target_ws, target_range, insert_mode, source_cells)
    
    # Calculate proper target range based on source dimensions
    logger.info(f"Original target_range: {target_range}")
    target_range = calculate_target_range(target_range, source_cells)
    logger.info(f"Calculated target_range: {target_range}")
    
    target_cells = target_ws[target_range]
    logger.info(f"Target cells type after calculation: {type(target_cells)}")
    
    # Ensure dimensions match for replace mode
    if insert_mode == "replace":
        if not check_dimensions_match(source_cells, target_cells):
            raise ValueError("Source and target range dimensions don't match for replace mode")
    
    # Copy data based on copy types
    copied_count = 0
    
    logger.debug(f"Source cells type: {type(source_cells)}")
    logger.debug(f"Target cells type: {type(target_cells)}")
    logger.debug(f"Source has coordinate: {hasattr(source_cells, 'coordinate')}")
    logger.debug(f"Target has coordinate: {hasattr(target_cells, 'coordinate')}")
    
    try:
        # Check if source_cells is a single Cell object
        if hasattr(source_cells, 'coordinate'):
            # Single cell
            logger.debug("Processing single cell")
            copy_cell(source_cells, target_cells, copy_types)
            copied_count += 1
        elif isinstance(source_cells, tuple) and len(source_cells) > 0 and isinstance(source_cells[0], tuple):
            # Multiple rows
            logger.debug(f"Processing multiple rows: {len(source_cells)} rows")
            for src_row, tgt_row in zip(source_cells, target_cells):
                if isinstance(src_row, tuple):
                    for src_cell, tgt_cell in zip(src_row, tgt_row):
                        copy_cell(src_cell, tgt_cell, copy_types)
                        copied_count += 1
                else:
                    copy_cell(src_row, tgt_row, copy_types)
                    copied_count += 1
        else:
            # Single row or multiple cells in one row
            logger.debug(f"Processing single row, source is tuple: {isinstance(source_cells, tuple)}")
            if isinstance(source_cells, tuple):
                for src_cell, tgt_cell in zip(source_cells, target_cells):
                    copy_cell(src_cell, tgt_cell, copy_types)
                    copied_count += 1
            else:
                copy_cell(source_cells, target_cells, copy_types)
                copied_count += 1
    except Exception as e:
        logger.error(f"Error during cell copying: {e}")
        logger.error(f"Source cells: {source_cells}")
        logger.error(f"Target cells: {target_cells}")
        raise
    
    logger.info(f"Copied {copied_count} cells")
    
    # Build result
    result = {"copied_cells": copied_count, "copy_types": copy_types}
    
    # Add insert_mode to result if it's not "replace"
    if insert_mode != "replace":
        result["insert_mode"] = insert_mode
    
    return result

def copy_formula(formula: str, target_ws: Worksheet, target_range: str,
                copy_types: List[str], insert_mode, logger: logging.Logger) -> dict:
    """Copy formula to target range."""
    if "formulas" not in copy_types:
        raise ValueError("Formula source requires 'formulas' in copy types")
    
    target_cells = target_ws[target_range]
    
    # Handle single cell vs range
    if hasattr(target_cells, 'coordinate'):
        # Single cell
        target_cells.value = formula
        count = 1
    else:
        # Multiple cells - set formula to first cell
        if isinstance(target_cells, tuple) and len(target_cells) > 0:
            first_cell = target_cells[0]
            if isinstance(first_cell, tuple):
                first_cell[0].value = formula
            else:
                first_cell.value = formula
        else:
            # Should not happen but handle gracefully
            target_cells.value = formula
        count = 1
    
    logger.info(f"Set formula '{formula}' to {count} cell(s)")
    return {"formula_set": count, "formula": formula}

def copy_value(value: str, target_ws: Worksheet, target_range: str,
              copy_types: List[str], insert_mode, logger: logging.Logger) -> dict:
    """Copy value to target range."""
    if "values" not in copy_types:
        raise ValueError("Value source requires 'values' in copy types")
    
    target_cells = target_ws[target_range]
    
    # Handle single cell vs range  
    if hasattr(target_cells, 'coordinate'):
        # Single cell
        target_cells.value = value
        count = 1
    else:
        # Multiple cells - set value to all cells
        count = 0
        if isinstance(target_cells, tuple) and len(target_cells) > 0 and isinstance(target_cells[0], tuple):
            # Multiple rows
            for row in target_cells:
                for cell in row:
                    cell.value = value
                    count += 1
        elif isinstance(target_cells, tuple):
            # Single row
            for cell in target_cells:
                cell.value = value
                count += 1
        else:
            # This shouldn't happen if we checked hasattr(coordinate) correctly
            target_cells.value = value
            count += 1
    
    logger.info(f"Set value '{value}' to {count} cell(s)")
    return {"values_set": count, "value": value}

def smart_convert_value(source_value, target_cell):
    """
    Smart value conversion with type detection and casting.
    
    Args:
        source_value: Value from source cell
        target_cell: Target cell to check existing type/format
        
    Returns:
        Converted value with appropriate type
    """
    if source_value is None:
        return None
        
    # If source is already correct type, return as is
    if isinstance(source_value, (int, float, bool)):
        return source_value
        
    # If source is string, try to detect and convert type
    if isinstance(source_value, str):
        # Empty string handling
        if not source_value.strip():
            return None
            
        # Formula handling - keep as string
        if source_value.startswith('='):
            return source_value
            
        # Try to detect numeric types
        source_clean = source_value.strip()
        
        # Boolean detection
        if source_clean.lower() in ('true', 'false', 'да', 'нет', 'yes', 'no'):
            return source_clean.lower() in ('true', 'да', 'yes')
            
        # Integer detection
        try:
            # Handle negative numbers
            if source_clean.lstrip('-').isdigit():
                return int(source_clean)
        except (ValueError, AttributeError):
            pass
            
        # Float detection
        try:
            # Try to convert to float
            float_val = float(source_clean.replace(',', '.'))  # Handle comma decimal separator
            # If it's actually an integer, return as int
            if float_val.is_integer():
                return int(float_val)
            return float_val
        except (ValueError, AttributeError):
            pass
            
        # Date/datetime detection (basic patterns)
        import re
        date_patterns = [
            r'^\d{4}-\d{2}-\d{2}$',  # YYYY-MM-DD
            r'^\d{2}/\d{2}/\d{4}$',  # MM/DD/YYYY or DD/MM/YYYY
            r'^\d{2}\.\d{2}\.\d{4}$'  # DD.MM.YYYY
        ]
        
        for pattern in date_patterns:
            if re.match(pattern, source_clean):
                try:
                    from datetime import datetime
                    if '-' in source_clean:
                        return datetime.strptime(source_clean, '%Y-%m-%d').date()
                    elif '/' in source_clean:
                        # Try both formats
                        try:
                            return datetime.strptime(source_clean, '%m/%d/%Y').date()
                        except ValueError:
                            return datetime.strptime(source_clean, '%d/%m/%Y').date()
                    elif '.' in source_clean:
                        return datetime.strptime(source_clean, '%d.%m.%Y').date()
                except (ValueError, ImportError):
                    pass
                    
    # If no conversion possible, return as string
    return source_value

def copy_cell(source_cell, target_cell, copy_types: List[str]):
    """Copy cell content based on copy types."""
    if "values" in copy_types:
        # Use smart value conversion instead of direct assignment
        converted_value = smart_convert_value(source_cell.value, target_cell)
        target_cell.value = converted_value
    
    if "formulas" in copy_types:
        # In openpyxl, formulas are stored in value, but we can check if it starts with =
        if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
            target_cell.value = source_cell.value
        elif "values" not in copy_types:
            # Only copy formula if values weren't already copied
            target_cell.value = source_cell.value
        
    if "formatting" in copy_types:
        # Copy formatting with proper handling of StyleProxy objects
        try:
            if source_cell.font:
                from copy import copy
                target_cell.font = copy(source_cell.font)
        except (AttributeError, TypeError):
            pass  # Skip font copying if it fails
            
        try:
            if source_cell.border:
                from copy import copy
                target_cell.border = copy(source_cell.border)
        except (AttributeError, TypeError):
            pass  # Skip border copying if it fails
            
        try:
            if source_cell.fill:
                from copy import copy
                target_cell.fill = copy(source_cell.fill)
        except (AttributeError, TypeError):
            pass  # Skip fill copying if it fails
            
        try:
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
        except (AttributeError, TypeError):
            pass  # Skip number format copying if it fails
            
        try:
            if source_cell.alignment:
                from copy import copy
                target_cell.alignment = copy(source_cell.alignment)
        except (AttributeError, TypeError):
            pass  # Skip alignment copying if it fails

def handle_insert_mode(ws: Worksheet, target_range: str, insert_mode, source_cells):
    """Handle insert mode by inserting rows/columns."""
    # Parse target range to get starting position
    range_obj = ws[target_range] 
    
    if hasattr(range_obj, 'coordinate'):
        # Single cell
        start_row, start_col = range_obj.row, range_obj.column
    else:
        # Multiple cells - get first cell position
        if isinstance(range_obj, tuple) and len(range_obj) > 0:
            first_cell = range_obj[0]
            if isinstance(first_cell, tuple):
                # Multiple rows - get first cell of first row
                start_row, start_col = first_cell[0].row, first_cell[0].column
            else:
                # Single row - get first cell
                start_row, start_col = first_cell.row, first_cell.column
        else:
            raise ValueError(f"Cannot parse target range: {target_range}")
    
    # Calculate dimensions needed
    if hasattr(source_cells, 'coordinate'):
        # Single cell
        rows_needed = 1
        cols_needed = 1
    elif isinstance(source_cells, tuple) and len(source_cells) > 0 and isinstance(source_cells[0], tuple):
        # Multiple rows
        rows_needed = len(source_cells)
        cols_needed = len(source_cells[0])
    else:
        # Single row or multiple cells in one row
        rows_needed = 1
        cols_needed = len(source_cells) if isinstance(source_cells, tuple) else 1
    
    # Insert rows if needed
    if "rows" in insert_mode:
        ws.insert_rows(start_row, rows_needed)
    
    # Insert columns if needed  
    if "columns" in insert_mode:
        ws.insert_cols(start_col, cols_needed)

def check_dimensions_match(source_cells, target_cells) -> bool:
    """Check if source and target cell ranges have matching dimensions."""
    # Get dimensions of source
    if hasattr(source_cells, 'coordinate'):
        # Single cell
        src_rows, src_cols = 1, 1
    elif isinstance(source_cells, tuple) and len(source_cells) > 0 and isinstance(source_cells[0], tuple):
        # Multiple rows
        src_rows, src_cols = len(source_cells), len(source_cells[0])
    else:
        # Single row or multiple cells in one row
        src_rows = 1
        src_cols = len(source_cells) if isinstance(source_cells, tuple) else 1
    
    # Get dimensions of target
    if hasattr(target_cells, 'coordinate'):
        # Single cell
        tgt_rows, tgt_cols = 1, 1
    elif isinstance(target_cells, tuple) and len(target_cells) > 0 and isinstance(target_cells[0], tuple):
        # Multiple rows
        tgt_rows, tgt_cols = len(target_cells), len(target_cells[0])
    else:
        # Single row or multiple cells in one row
        tgt_rows = 1
        tgt_cols = len(target_cells) if isinstance(target_cells, tuple) else 1
        
    return src_rows == tgt_rows and src_cols == tgt_cols

def analyze_workspace_files(workspace: dict, context: dict = None) -> dict:
    """Analyze workspace files and return statistics about data ranges."""
    try:
        file_handlers = {}
        results = {}
        
        for file_id, file_path in workspace.items():
            try:
                # Load the file
                if file_path.lower().endswith('.csv'):
                    wb = load_csv_as_excel(file_path)
                    sheet_name = "Sheet1"  # CSV becomes Sheet1
                else:
                    wb = load_workbook(file_path, file_handlers)
                
                file_stats = {
                    "file_path": file_path,
                    "file_type": "csv" if file_path.lower().endswith('.csv') else "excel",
                    "sheets": {}
                }
                
                # Analyze each sheet
                for sheet in wb.worksheets:
                    sheet_stats = analyze_sheet_data(sheet)
                    file_stats["sheets"][sheet.title] = sheet_stats
                
                results[file_id] = file_stats
                
            except Exception as e:
                results[file_id] = {
                    "error": f"Could not analyze file: {str(e)}",
                    "file_path": file_path
                }
        
        return {
            "success": True,
            "analysis": results
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": f"Analysis failed: {str(e)}"
        }

def analyze_sheet_data(sheet) -> dict:
    """Analyze a single sheet and return data statistics."""
    max_row = 1
    max_col = 1
    filled_cells = 0
    
    # Find the actual data boundaries
    for row_num, row in enumerate(sheet.iter_rows(), 1):
        row_has_data = False
        for col_num, cell in enumerate(row, 1):
            if cell.value is not None and str(cell.value).strip() != "":
                filled_cells += 1
                max_col = max(max_col, col_num)
                row_has_data = True
        if row_has_data:
            max_row = max(max_row, row_num)
    
    # Convert column number to letter (A, B, C, ..., AA, AB, etc.)
    def col_num_to_letter(col_num):
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + ord('A')) + result
            col_num //= 26
        return result
    
    return {
        "max_row": max_row,
        "max_col": max_col,
        "max_col_letter": col_num_to_letter(max_col),
        "filled_cells": filled_cells,
        "data_range": f"A1:{col_num_to_letter(max_col)}{max_row}",
        "suggested_range": f"A2:{col_num_to_letter(max_col)}{max_row}"  # Skip header
    }


def save_workbook(file_path: str, wb, file_handlers: dict):
    """Save workbook to file."""
    path = Path(file_path)
    
    if path.suffix.lower() == '.csv':
        # Save as CSV
        save_excel_as_csv(wb, file_path)
    else:
        # Force formula recalculation before saving Excel
        # Save as Excel
        wb.save(file_path)

def save_excel_as_csv(wb, csv_path: str):
    """Save Excel workbook as CSV (first sheet only)."""
    ws = wb.active
    
    # Convert to pandas DataFrame
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    if data:
        df = pd.DataFrame(data[1:], columns=data[0])  # First row as headers
        df.to_csv(csv_path, index=False)

def save_save_as_files(save_as_files: dict, workspace: dict, logger: logging.Logger) -> dict:
    """Save all save_as files."""
    saved_files = []
    
    for save_as_id, file_info in save_as_files.items():
        try:
            wb = file_info['workbook']
            file_path = file_info['file_path']
            
            path = Path(file_path)
            if path.suffix.lower() == '.csv':
                # Save as CSV
                save_excel_as_csv(wb, file_path)
            else:
                # Force formula recalculation before saving Excel
                # Save as Excel
                wb.save(file_path)
                
            saved_files.append(file_path)
            logger.info(f"Saved save_as file '{save_as_id}': {file_path}")
        except Exception as e:
            logger.error(f"Failed to save save_as file '{save_as_id}': {str(e)}")
            
    return {"saved_files": saved_files}

def save_all_files(file_handlers: dict, logger: logging.Logger) -> dict:
    """Save all loaded files."""
    saved_files = []
    
    for file_path, wb in file_handlers.items():
        try:
            save_workbook(file_path, wb, file_handlers)
            saved_files.append(file_path)
            logger.info(f"Saved file: {file_path}")
        except Exception as e:
            logger.error(f"Failed to save {file_path}: {str(e)}")
            
    return {"saved_files": saved_files}
