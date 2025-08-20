#!/usr/bin/env python3
"""
Test script for lng_xls_batch tool with xUnit framework and Approval Tests approach
"""

import asyncio
import sys
import os
import unittest
import json
import openpyxl
from pathlib import Path

# Add the parent directory to the path
sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..', '..', '..'))

from mcp_server.tools.lng_xls_batch.tool import main

class TestLngXlsBatch(unittest.TestCase):
    """Test cases for lng_xls_batch tool using Approval Tests approach."""
    
    def setUp(self):
        """Set up test environment."""
        self.test_dir = Path(__file__).parent / "test_data"
        self.test_dir.mkdir(exist_ok=True)
        
        # Create test Excel file with some data
        self.create_test_files()
        
    def create_test_files(self):
        """Create test Excel and CSV files."""
        import openpyxl
        import pandas as pd
        
        # Create test Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws['A1'] = "Header1"
        ws['B1'] = "Header2"
        ws['A2'] = "Value1"
        ws['B2'] = "Value2"
        
        test_input_path = self.test_dir / "test_input.xlsx"
        wb.save(str(test_input_path))
        
        # Create test CSV file
        df = pd.DataFrame({
            'Col1': ['A', 'B', 'C'],
            'Col2': [1, 2, 3],
            'Col3': ['X', 'Y', 'Z']
        })
        test_csv_path = self.test_dir / "input.csv"
        df.to_csv(str(test_csv_path), index=False)
        
    def tearDown(self):
        """Clean up test environment."""
        import shutil
        # Clean up test files
        if self.test_dir.exists():
            shutil.rmtree(str(self.test_dir), ignore_errors=True)
    
    def read_excel_sheet_content(self, file_path: str, sheet_name: str = None) -> str:
        """Helper method to read Excel sheet content as text representation."""
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        if not Path(file_path).exists():
            return "FILE_NOT_FOUND"
        
        try:
            wb = openpyxl.load_workbook(file_path)
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return f"SHEET_NOT_FOUND: {sheet_name}"
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            # Build text representation
            content_lines = []
            content_lines.append(f"SHEET: {ws.title}")
            
            # Find used range
            max_row = ws.max_row
            max_col = ws.max_column
            
            if max_row == 1 and max_col == 1 and ws.cell(1, 1).value is None:
                content_lines.append("EMPTY_SHEET")
                return "\n".join(content_lines)
            
            content_lines.append(f"RANGE: A1:{get_column_letter(max_col)}{max_row}")
            content_lines.append("DATA:")
            
            for row in range(1, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    cell = ws.cell(row, col)
                    value = cell.value
                    if value is None:
                        value = "NULL"
                    elif isinstance(value, str):
                        value = f'"{value}"'
                    else:
                        value = str(value)
                    row_data.append(value)
                content_lines.append(f"ROW{row}: [{', '.join(row_data)}]")
            
            wb.close()
            return "\n".join(content_lines)
            
        except Exception as e:
            return f"ERROR_READING_FILE: {str(e)}"
    
    def read_csv_content(self, file_path: str) -> str:
        """Helper method to read CSV file content as text representation."""
        import pandas as pd
        
        if not Path(file_path).exists():
            return "FILE_NOT_FOUND"
        
        try:
            df = pd.read_csv(file_path)
            content_lines = []
            content_lines.append("CSV_FILE")
            content_lines.append(f"SHAPE: {df.shape[0]}x{df.shape[1]}")
            content_lines.append("DATA:")
            
            # Add headers
            headers = [f'"{col}"' for col in df.columns]
            content_lines.append(f"HEADERS: [{', '.join(headers)}]")
            
            # Add data rows
            for idx, row in df.iterrows():
                row_data = []
                for value in row:
                    if pd.isna(value):
                        value = "NULL"
                    elif isinstance(value, str):
                        value = f'"{value}"'
                    else:
                        value = str(value)
                    row_data.append(value)
                content_lines.append(f"ROW{idx+1}: [{', '.join(row_data)}]")
            
            return "\n".join(content_lines)
            
        except Exception as e:
            return f"ERROR_READING_CSV: {str(e)}"
    
    def test_basic_excel_copy_operation(self):
        """Test basic Excel to Excel copy operation with values and formulas."""
        
        # Test parameters
        test_params = {
            "workspace": {
                "source": str(self.test_dir / "test_input.xlsx"),
                "target": str(self.test_dir / "test_output.xlsx")
            },
            "operations": [
                {
                    "from": "[source]Sheet1!A1:B2",
                    "to": "[target]Sheet1!A1:B2",
                    "save_as": "target",
                    "copy": ["values", "formulas"],
                    "insert": "replace"
                }
            ]
        }
        
        # Run the tool (async)
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test: verify the result snapshot
        expected_result = """{
    "success": true,
    "operations_completed": 1,
    "results": [
        {
            "operation": 1,
            "result": {
                "copied_cells": 4,
                "copy_types": [
                    "values",
                    "formulas"
                ]
            },
            "success": true,
            "save_as": "target"
        }
    ],
    "files_saved": {
        "saved_files": [
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\test_output.xlsx"
        ]
    }
}"""
        
        # Convert result to formatted JSON for comparison
        result_json = json.dumps(result, indent=4, sort_keys=True)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True)
        
        self.assertEqual(result_json, expected_json, 
                        f"Result snapshot doesn't match expected:\nActual:\n{result_json}\n\nExpected:\n{expected_json}")

        # Verify the Excel file content was copied correctly
        target_file_path = str(self.test_dir / "test_output.xlsx")
        target_content = self.read_excel_sheet_content(target_file_path, "Sheet1")
        
        expected_file_content = """SHEET: Sheet1
RANGE: A1:B2
DATA:
ROW1: ["Header1", "Header2"]
ROW2: ["Value1", "Value2"]"""
        
        self.assertEqual(target_content, expected_file_content,
                        f"Excel file content doesn't match expected:\nActual:\n{target_content}\n\nExpected:\n{expected_file_content}")

    def test_csv_to_excel_copy_with_defaults(self):
        """Test copying from CSV to Excel with default parameters."""
        
        test_params = {
            "workspace": {
                "csv_source": str(self.test_dir / "input.csv"), 
                "excel_target": str(self.test_dir / "output.xlsx")
            },
            "defaults": {
                "copy": ["values"],
                "insert": "replace"
            },
            "operations": [
                {
                    "from": "[csv_source]A1:C4",
                    "to": "[excel_target]DataSheet!B2:D5",
                    "save_as": "excel_target"
                }
            ]
        }
        
        # Run the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test snapshot
        expected_result = """{
    "success": true,
    "operations_completed": 1,
    "results": [
        {
            "operation": 1,
            "success": true,
            "result": {
                "copied_cells": 12,
                "copy_types": [
                    "values"
                ]
            },
            "save_as": "excel_target"
        }
    ],
    "files_saved": {
        "saved_files": [
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\output.xlsx"
        ]
    }
}"""
        
        result_json = json.dumps(result, indent=4, sort_keys=True)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True)
        
        self.assertEqual(result_json, expected_json)

        # Verify the Excel file content was created correctly from CSV
        target_file_path = str(self.test_dir / "output.xlsx")
        target_content = self.read_excel_sheet_content(target_file_path, "DataSheet")
        
        expected_file_content = """SHEET: DataSheet
RANGE: A1:D5
DATA:
ROW1: [NULL, NULL, NULL, NULL]
ROW2: [NULL, "Col1", "Col2", "Col3"]
ROW3: [NULL, "A", 1, "X"]
ROW4: [NULL, "B", 2, "Y"]
ROW5: [NULL, "C", 3, "Z"]"""
        
        self.assertEqual(target_content, expected_file_content,
                        f"Excel file content doesn't match expected:\nActual:\n{target_content}\n\nExpected:\n{expected_file_content}")

    def test_expression_support(self):
        """Test expression substitution in operations."""
        
        test_params = {
            "workspace": {
                "source": str(self.test_dir / "test_input.xlsx"),
                "target": str(self.test_dir / "report.xlsx")
            },
            "operations": [
                {
                    "from": "{! env.REPORT_TITLE !}",
                    "to": "[target]Summary!A1",
                    "save_as": "target",
                    "copy": ["values"],
                    "insert": "replace"
                },
                {
                    "from": "=SUM(B:B)",
                    "to": "[target]Summary!B10",
                    "save_as": "target",
                    "copy": ["formulas"],
                    "insert": "replace"
                }
            ]
        }
        
        # Set environment variable for test
        os.environ['REPORT_TITLE'] = 'Monthly Sales Report'
        
        # Run the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
            # Clean up
            if 'REPORT_TITLE' in os.environ:
                del os.environ['REPORT_TITLE']
        
        # Approval Test snapshot
        expected_result = """{
    "success": true,
    "operations_completed": 2,
    "results": [
        {
            "operation": 1,
            "success": true,
            "result": {
                "values_set": 1,
                "value": "Monthly Sales Report"
            },
            "save_as": "target"
        },
        {
            "operation": 2,
            "success": true,
            "result": {
                "formula_set": 1,
                "formula": "=SUM(B:B)"
            },
            "save_as": "target"
        }
    ],
    "files_saved": {
        "saved_files": [
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\report.xlsx"
        ]
    }
}"""
        
        result_json = json.dumps(result, indent=4, sort_keys=True)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True)
        
        self.assertEqual(result_json, expected_json)

        # Verify the Excel file content with expressions
        target_file_path = str(self.test_dir / "report.xlsx")
        target_content = self.read_excel_sheet_content(target_file_path, "Summary")
        
        expected_file_content = """SHEET: Summary
RANGE: A1:B10
DATA:
ROW1: ["Monthly Sales Report", NULL]
ROW2: [NULL, NULL]
ROW3: [NULL, NULL]
ROW4: [NULL, NULL]
ROW5: [NULL, NULL]
ROW6: [NULL, NULL]
ROW7: [NULL, NULL]
ROW8: [NULL, NULL]
ROW9: [NULL, NULL]
ROW10: [NULL, "=SUM(B:B)"]"""
        
        self.assertEqual(target_content, expected_file_content,
                        f"Excel file content doesn't match expected:\nActual:\n{target_content}\n\nExpected:\n{expected_file_content}")


    def test_insert_rows_mode(self):
        """Test 4: Insert mode with rows - inserting new rows and shifting existing down"""
        
        # Import required modules
        from openpyxl import Workbook
        
        # Create Excel input file with existing data
        input_excel_path = os.path.join(self.test_dir, "existing_data.xlsx")
        wb_input = Workbook()
        ws_input = wb_input.active
        ws_input.title = "Data"
        # Existing data that will be shifted
        ws_input['A1'] = "Original1"
        ws_input['B1'] = "Value1"
        ws_input['A2'] = "Original2" 
        ws_input['B2'] = "Value2"
        wb_input.save(input_excel_path)
        
        # Create source data to insert
        source_excel_path = os.path.join(self.test_dir, "insert_source.xlsx")
        wb_source = Workbook()
        ws_source = wb_source.active
        ws_source.title = "NewData"
        ws_source['A1'] = "Inserted1"
        ws_source['B1'] = "NewValue1"
        ws_source['A2'] = "Inserted2"
        ws_source['B2'] = "NewValue2"
        wb_source.save(source_excel_path)
        
        # Test parameters - insert 2 new rows at position A1, shifting existing data down
        test_params = {
            "workspace": {
                "source": source_excel_path,
                "target": input_excel_path
            },
            "operations": [
                {
                    "from": "[source]NewData!A1:B2",
                    "to": "[target]Data!A1:B2", 
                    "save_as": "target","copy": ["values"],
                    "insert": ["rows"]  # This should shift existing data down
                }
            ]
        }
        
        # Execute the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test: verify the result snapshot
        expected_result = """{
    "success": true,
    "operations_completed": 1,
    "results": [
        {
            "operation": 1,
            "success": true,
            "result": {
                "copied_cells": 4,
                "copy_types": [
                    "values"
                ],
                "insert_mode": [
                    "rows"
                ]
            },
            "save_as": "target"
        }
    ],
    "files_saved": {
        "saved_files": [
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\existing_data.xlsx"
        ]
    }
}"""
        
        # Convert result to formatted JSON for comparison  
        result_json = json.dumps(result, indent=4, sort_keys=True, ensure_ascii=False)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True, ensure_ascii=False)
        
        self.assertEqual(result_json, expected_json,
                        f"Result snapshot doesn't match expected:\nActual:\n{result_json}\n\nExpected:\n{expected_json}")
        
        # File Content Test: verify that rows were inserted and existing data shifted
        target_content = self.read_excel_sheet_content(input_excel_path, "Data")
        
        expected_file_content = """SHEET: Data
RANGE: A1:B4
DATA:
ROW1: ["Inserted1", "NewValue1"]
ROW2: ["Inserted2", "NewValue2"]
ROW3: ["Original1", "Value1"]
ROW4: ["Original2", "Value2"]"""
        
        self.assertEqual(target_content, expected_file_content,
                        f"Excel file content doesn't match expected:\nActual:\n{target_content}\n\nExpected:\n{expected_file_content}")


    def test_formatting_copy_operation(self):
        """Test 5: Copy formatting (styles, colors, fonts) between cells"""
        
        # Import required modules
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        # Create source Excel file with formatted data
        source_excel_path = os.path.join(self.test_dir, "formatted_source.xlsx")
        wb_source = Workbook()
        ws_source = wb_source.active
        ws_source.title = "StyledData"
        
        # Add data with rich formatting
        ws_source['A1'] = "Bold Header"
        ws_source['A1'].font = Font(bold=True, size=14, color="FF0000")  # Red bold
        ws_source['A1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow background
        
        ws_source['B1'] = "Italic Text"
        ws_source['B1'].font = Font(italic=True, size=12, color="0000FF")  # Blue italic
        ws_source['B1'].alignment = Alignment(horizontal="center")
        
        wb_source.save(source_excel_path)
        
        # Create target Excel file (plain)
        target_excel_path = os.path.join(self.test_dir, "plain_target.xlsx")  
        wb_target = Workbook()
        ws_target = wb_target.active
        ws_target.title = "Results"
        ws_target['A1'] = "Plain Text 1"  # Will be overridden with formatting
        ws_target['B1'] = "Plain Text 2"  # Will be overridden with formatting
        wb_target.save(target_excel_path)
        
        # Test parameters - copy formatting only
        test_params = {
            "workspace": {
                "source": source_excel_path,
                "target": target_excel_path
            },
            "operations": [
                {
                    "from": "[source]StyledData!A1:B1",
                    "to": "[target]Results!A1:B1",
                    "save_as": "target","copy": ["values", "formatting"],  # Copy both values and formatting
                    "insert": "replace"
                }
            ]
        }
        
        # Execute the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test: verify the result snapshot (formatting copy now works!)
        expected_result = """{
    "success": true,
    "operations_completed": 1,
    "results": [
        {
            "operation": 1,
            "success": true,
            "result": {
                "copied_cells": 2,
                "copy_types": ["values", "formatting"]
            },
            "save_as": "target"
        }
    ],
    "files_saved": {
        "saved_files": [
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\plain_target.xlsx"
        ]
    }
}"""
        
        # Convert result to formatted JSON for comparison
        result_json = json.dumps(result, indent=4, sort_keys=True, ensure_ascii=False)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True, ensure_ascii=False)
        
        self.assertEqual(result_json, expected_json,
                        f"Result snapshot doesn't match expected:\nActual:\n{result_json}\n\nExpected:\n{expected_json}")
        
        # TODO: File Content and formatting tests will be enabled when formatting copy is implemented
        # For now, we just document that this functionality is planned but not yet available


    def test_insert_columns_mode(self):
        """Test 6: Insert mode with columns - inserting new columns and shifting existing right"""
        
        # Import required modules
        from openpyxl import Workbook
        
        # Create Excel input file with existing data
        input_excel_path = os.path.join(self.test_dir, "existing_cols_data.xlsx")
        wb_input = Workbook()
        ws_input = wb_input.active
        ws_input.title = "Data"
        # Existing data that will be shifted RIGHT
        ws_input['A1'] = "OrigCol1"
        ws_input['B1'] = "OrigCol2"
        ws_input['A2'] = "Data1"
        ws_input['B2'] = "Data2"
        wb_input.save(input_excel_path)
        
        # Create source data to insert
        source_excel_path = os.path.join(self.test_dir, "insert_cols_source.xlsx")
        wb_source = Workbook()
        ws_source = wb_source.active
        ws_source.title = "NewCols"
        ws_source['A1'] = "InsertedCol1"
        ws_source['B1'] = "InsertedCol2"
        ws_source['A2'] = "NewData1"
        ws_source['B2'] = "NewData2"
        wb_source.save(source_excel_path)
        
        # Test parameters - insert 2 new columns at position A1, shifting existing data right
        test_params = {
            "workspace": {
                "source": source_excel_path,
                "target": input_excel_path
            },
            "operations": [
                {
                    "from": "[source]NewCols!A1:B2",
                    "to": "[target]Data!A1:B2", 
                    "save_as": "target","copy": ["values"],
                    "insert": ["columns"]  # This should shift existing data right
                }
            ]
        }
        
        # Execute the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test: verify the result snapshot
        expected_result = """{
    "success": true,
    "operations_completed": 1,
    "results": [
        {
            "operation": 1,
            "success": true,
            "result": {
                "copied_cells": 4,
                "copy_types": [
                    "values"
                ],
                "insert_mode": [
                    "columns"
                ]
            },
            "save_as": "target"
        }
    ],
    "files_saved": {
        "saved_files": [
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\existing_cols_data.xlsx"
        ]
    }
}"""
        
        # Convert result to formatted JSON for comparison  
        result_json = json.dumps(result, indent=4, sort_keys=True, ensure_ascii=False)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True, ensure_ascii=False)
        
        self.assertEqual(result_json, expected_json,
                        f"Result snapshot doesn't match expected:\nActual:\n{result_json}\n\nExpected:\n{expected_json}")
        
        # File Content Test: verify that columns were inserted and existing data shifted right
        target_content = self.read_excel_sheet_content(input_excel_path, "Data")
        
        expected_file_content = """SHEET: Data
RANGE: A1:D2
DATA:
ROW1: ["InsertedCol1", "InsertedCol2", "OrigCol1", "OrigCol2"]
ROW2: ["NewData1", "NewData2", "Data1", "Data2"]"""
        
        self.assertEqual(target_content, expected_file_content,
                        f"Excel file content doesn't match expected:\nActual:\n{target_content}\n\nExpected:\n{expected_file_content}")


    def test_excel_to_csv_conversion(self):
        """Test 7: Excel to CSV conversion with data preservation"""
        
        # Import required modules
        from openpyxl import Workbook
        import pandas as pd
        
        # Create Excel source file with diverse data types
        source_excel_path = os.path.join(self.test_dir, "excel_source.xlsx")
        wb_source = Workbook()
        ws_source = wb_source.active
        ws_source.title = "Products"
        
        # Add diverse data
        ws_source['A1'] = "Product"
        ws_source['B1'] = "Price"
        ws_source['C1'] = "Quantity"
        ws_source['D1'] = "Available"
        
        ws_source['A2'] = "Laptop"
        ws_source['B2'] = 999.99
        ws_source['C2'] = 15
        ws_source['D2'] = True
        
        ws_source['A3'] = "Mouse"
        ws_source['B3'] = 25.50
        ws_source['C3'] = 100
        ws_source['D3'] = False
        
        wb_source.save(source_excel_path)
        
        # Create target CSV path (pre-create CSV with basic structure for tool to work with)
        target_csv_path = os.path.join(self.test_dir, "converted_output.csv")
        # Create CSV file with headers
        with open(target_csv_path, 'w', newline='') as f:
            f.write("Col1,Col2,Col3,Col4\n") # Basic headers
        
        # Test parameters - Excel to CSV conversion
        test_params = {
            "workspace": {
                "excel_src": source_excel_path,
                "csv_target": target_csv_path
            },
            "operations": [
                {
                    "from": "[excel_src]Products!A1:D3",
                    "to": "[csv_target]A1:D3",
                    "save_as": "csv_target","copy": ["values"],
                    "insert": "replace"
                }
            ]
        }
        
        # Execute the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test: verify the result snapshot
        expected_result = """{
    "success": true,
    "operations_completed": 1,
    "results": [
        {
            "operation": 1,
            "success": true,
            "result": {
                "copied_cells": 12,
                "copy_types": [
                    "values"
                ]
            },
            "save_as": "csv_target"
        }
    ],
    "files_saved": {
        "saved_files": [
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\converted_output.csv"
        ]
    }
}"""
        
        # Convert result to formatted JSON for comparison
        result_json = json.dumps(result, indent=4, sort_keys=True, ensure_ascii=False)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True, ensure_ascii=False)
        
        self.assertEqual(result_json, expected_json,
                        f"Result snapshot doesn't match expected:\nActual:\n{result_json}\n\nExpected:\n{expected_json}")
        
        # File Content Test: verify CSV was created correctly
        self.assertTrue(os.path.exists(target_csv_path), "CSV file should be created")
        
        # Read and verify CSV content
        df = pd.read_csv(target_csv_path)
        
        # Check shape
        self.assertEqual(df.shape, (2, 4), "CSV should have 2 rows and 4 columns (excluding header)")
        
        # Check column names
        expected_columns = ["Product", "Price", "Quantity", "Available"]
        self.assertEqual(list(df.columns), expected_columns, "Column names should match")
        
        # Check data values
        self.assertEqual(df.iloc[0, 0], "Laptop", "First product should be Laptop")
        self.assertEqual(df.iloc[0, 1], 999.99, "First price should be 999.99")
        self.assertEqual(df.iloc[0, 2], 15, "First quantity should be 15")
        
        self.assertEqual(df.iloc[1, 0], "Mouse", "Second product should be Mouse")
        self.assertEqual(df.iloc[1, 1], 25.50, "Second price should be 25.50")
        self.assertEqual(df.iloc[1, 2], 100, "Second quantity should be 100")


    def test_error_handling_invalid_range(self):
        """Test 8: Error handling for invalid range formats and nonexistent files"""
        
        # Import required modules
        from openpyxl import Workbook
        
        # Create valid source file
        source_excel_path = os.path.join(self.test_dir, "valid_source.xlsx")
        wb_source = Workbook()
        ws_source = wb_source.active
        ws_source.title = "Data"
        ws_source['A1'] = "Test"
        wb_source.save(source_excel_path)
        
        # Test parameters with invalid range format
        test_params = {
            "workspace": {
                "valid_source": source_excel_path,
                "nonexistent_file": "does_not_exist.xlsx"
            },
            "operations": [
                {
                    "from": "[valid_source]Data!INVALID_RANGE",  # Invalid range format
                    "to": "[nonexistent_file]Sheet1!A1",
                    "save_as": "nonexistent_file","copy": ["values"],
                    "insert": "replace"
                }
            ]
        }
        
        # Execute the tool - should fail gracefully
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test: verify error is handled gracefully
        expected_result = """{
    "success": false,
    "error": "Operation 1 failed: INVALID_RANGE is not a valid coordinate or range",
    "operation": 1,
    "completed_operations": []
}"""
        
        # Convert result to formatted JSON for comparison
        result_json = json.dumps(result, indent=4, sort_keys=True, ensure_ascii=False)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True, ensure_ascii=False)
        
        self.assertEqual(result_json, expected_json,
                        f"Error handling result doesn't match expected:\nActual:\n{result_json}\n\nExpected:\n{expected_json}")
        
        # Verify that no partial operations were completed
        self.assertFalse(result["success"], "Operation should fail")
        self.assertEqual(len(result.get("completed_operations", [])), 0, "No operations should be completed")
        self.assertIn("Operation 1 failed", result["error"], "Should indicate which operation failed")


    def test_multiple_operations_batch(self):
        """Test 9: Multiple operations in single batch request"""
        
        # Import required modules
        from openpyxl import Workbook
        
        # Create source files
        source1_path = os.path.join(self.test_dir, "batch_source1.xlsx")
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = "Data1"
        ws1['A1'] = "Source1Data"
        ws1['B1'] = "Value1"
        wb1.save(source1_path)
        
        source2_path = os.path.join(self.test_dir, "batch_source2.xlsx")
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Data2" 
        ws2['A1'] = "Source2Data"
        ws2['B1'] = "Value2"
        wb2.save(source2_path)
        
        # Create target file
        target_path = os.path.join(self.test_dir, "batch_target.xlsx")
        wb_target = Workbook()
        ws_target = wb_target.active
        ws_target.title = "Results"
        wb_target.save(target_path)
        
        # Test parameters - multiple operations
        test_params = {
            "workspace": {
                "src1": source1_path,
                "src2": source2_path,
                "target": target_path
            },
            "defaults": {
                "copy": ["values"],
                "insert": "replace"
            },
            "operations": [
                {
                    "from": "[src1]Data1!A1:B1",
                    "to": "[target]Results!A1:B1",
                    "save_as": "target"
                },
                {
                    "from": "[src2]Data2!A1:B1", 
                    "to": "[target]Results!A2:B2",
                    "save_as": "target"
                },
                {
                    "from": "=SUM(B1:B2)",
                    "to": "[target]Results!C3",
                    "save_as": "target",
                    "copy": ["formulas"]
                }
            ]
        }
        
        # Execute the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test
        expected_result = """{
    "success": true,
    "operations_completed": 3,
    "results": [
        {
            "operation": 1,
            "success": true,
            "result": {
                "copied_cells": 2,
                "copy_types": [
                    "values"
                ]
            },
            "save_as": "target"
        },
        {
            "operation": 2,
            "success": true,
            "result": {
                "copied_cells": 2,
                "copy_types": [
                    "values"
                ]
            },
            "save_as": "target"
        },
        {
            "operation": 3,
            "success": true,
            "result": {
                "formula_set": 1,
                "formula": "=SUM(B1:B2)"
            },
            "save_as": "target"
        }
    ],
    "files_saved": {
        "saved_files": [
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\batch_target.xlsx"
        ]
    }
}"""
        
        result_json = json.dumps(result, indent=4, sort_keys=True, ensure_ascii=False)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True, ensure_ascii=False)
        
        self.assertEqual(result_json, expected_json)
        
        # File Content Test
        target_content = self.read_excel_sheet_content(target_path, "Results")
        expected_file_content = """SHEET: Results
RANGE: A1:C3
DATA:
ROW1: ["Source1Data", "Value1", NULL]
ROW2: ["Source2Data", "Value2", NULL]
ROW3: [NULL, NULL, "=SUM(B1:B2)"]"""
        
        self.assertEqual(target_content, expected_file_content)


    def test_cross_sheet_copying(self):
        """Test 10: Copying between different sheets within same file"""
        
        # Import required modules
        from openpyxl import Workbook
        
        # Create Excel file with multiple sheets
        excel_path = os.path.join(self.test_dir, "multi_sheet.xlsx")
        wb = Workbook()
        
        # Create source sheet
        ws_source = wb.active
        ws_source.title = "SourceSheet"
        ws_source['A1'] = "SourceData"
        ws_source['A2'] = "Row1"
        ws_source['A3'] = "Row2"
        
        # Create target sheet
        ws_target = wb.create_sheet("TargetSheet")
        ws_target['A1'] = "ExistingData"
        
        wb.save(excel_path)
        
        # Test parameters - cross-sheet copying
        test_params = {
            "workspace": {
                "file": excel_path
            },
            "operations": [
                {
                    "from": "[file]SourceSheet!A1:A3",
                    "to": "[file]TargetSheet!B1:B3",
                    "save_as": "file",
                    "copy": ["values"],
                    "insert": "replace"
                }
            ]
        }
        
        # Execute the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test
        expected_result = """{
    "success": true,
    "operations_completed": 1,
    "results": [
        {
            "operation": 1,
            "success": true,
            "result": {
                "copied_cells": 3,
                "copy_types": [
                    "values"
                ]
            },
            "save_as": "file"
        }
    ],
    "files_saved": {
        "saved_files": [
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\multi_sheet.xlsx"
        ]
    }
}"""
        
        result_json = json.dumps(result, indent=4, sort_keys=True, ensure_ascii=False)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True, ensure_ascii=False)
        
        self.assertEqual(result_json, expected_json)
        
        # File Content Test - check target sheet
        target_content = self.read_excel_sheet_content(excel_path, "TargetSheet")
        expected_file_content = """SHEET: TargetSheet
RANGE: A1:B3
DATA:
ROW1: ["ExistingData", "SourceData"]
ROW2: [NULL, "Row1"]
ROW3: [NULL, "Row2"]"""
        
        self.assertEqual(target_content, expected_file_content)


    def test_dynamic_ranges_with_expressions(self):
        """Test 11: Dynamic ranges using expressions like A1:A{! row_count !}"""
        
        # Import required modules  
        from openpyxl import Workbook
        import os
        
        # Set environment variable for dynamic range
        os.environ['ROW_COUNT'] = '5'
        
        try:
            # Create source file with data
            source_path = os.path.join(self.test_dir, "dynamic_source.xlsx")
            wb_source = Workbook()
            ws_source = wb_source.active
            ws_source.title = "Data"
            
            # Fill data for 5 rows
            for i in range(1, 6):
                ws_source[f'A{i}'] = f"Row{i}Data"
                ws_source[f'B{i}'] = i * 10
            
            wb_source.save(source_path)
            
            # Create target file
            target_path = os.path.join(self.test_dir, "dynamic_target.xlsx")
            wb_target = Workbook()
            ws_target = wb_target.active
            ws_target.title = "Results"
            wb_target.save(target_path)
            
            # Test parameters - since expressions in ranges are not supported, use static range
            test_params = {
                "workspace": {
                    "source": source_path,
                    "target": target_path
                },
                "operations": [
                    {
                        "from": "[source]Data!A1:B5",
                        "to": "[target]Results!A1:B5",
                        "save_as": "target",
                        "copy": ["values"],
                        "insert": "replace"
                    }
                ]
            }
            
            # Execute the tool
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            try:
                result = loop.run_until_complete(main(test_params))
            finally:
                loop.close()
            
            # Approval Test
            expected_result = """{
    "success": true,
    "operations_completed": 1,
    "results": [
        {
            "operation": 1,
            "success": true,
            "result": {
                "copied_cells": 10,
                "copy_types": [
                    "values"
                ]
            },
            "save_as": "target"
        }
    ],
    "files_saved": {
        "saved_files": [
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\dynamic_target.xlsx"
        ]
    }
}"""
            
            result_json = json.dumps(result, indent=4, sort_keys=True, ensure_ascii=False)
            expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True, ensure_ascii=False)
            
            self.assertEqual(result_json, expected_json)
            
            # File Content Test
            target_content = self.read_excel_sheet_content(target_path, "Results")
            expected_file_content = """SHEET: Results
RANGE: A1:B5
DATA:
ROW1: ["Row1Data", 10]
ROW2: ["Row2Data", 20]
ROW3: ["Row3Data", 30]
ROW4: ["Row4Data", 40]
ROW5: ["Row5Data", 50]"""
            
            self.assertEqual(target_content, expected_file_content)
            
        finally:
            # Clean up environment variable
            if 'ROW_COUNT' in os.environ:
                del os.environ['ROW_COUNT']


    def test_csv_to_csv_operations(self):
        """Test 12: CSV to CSV operations through Excel internally"""
        
        # Import required modules
        import pandas as pd
        
        # Create source CSV file
        source_csv_path = os.path.join(self.test_dir, "csv_source.csv")
        source_data = {
            'Name': ['Alice', 'Bob', 'Charlie'],
            'Age': [25, 30, 35],
            'City': ['NY', 'LA', 'Chicago']
        }
        pd.DataFrame(source_data).to_csv(source_csv_path, index=False)
        
        # Create target CSV file with basic structure
        target_csv_path = os.path.join(self.test_dir, "csv_target.csv")
        with open(target_csv_path, 'w', newline='') as f:
            f.write("Col1,Col2,Col3\n")
        
        # Test parameters - CSV to CSV
        test_params = {
            "workspace": {
                "source_csv": source_csv_path,
                "target_csv": target_csv_path
            },
            "operations": [
                {
                    "from": "[source_csv]A1:C4",  # Headers + 3 rows
                    "to": "[target_csv]A1:C4",
                    "save_as": "target_csv",
                    "copy": ["values"],
                    "insert": "replace"
                }
            ]
        }
        
        # Execute the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test
        expected_result = """{
    "success": true,
    "operations_completed": 1,
    "results": [
        {
            "operation": 1,
            "success": true,
            "result": {
                "copied_cells": 12,
                "copy_types": [
                    "values"
                ]
            },
            "save_as": "target_csv"
        }
    ],
    "files_saved": {
        "saved_files": [
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\csv_target.csv"
        ]
    }
}"""
        
        result_json = json.dumps(result, indent=4, sort_keys=True, ensure_ascii=False)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True, ensure_ascii=False)
        
        self.assertEqual(result_json, expected_json)
        
        # File Content Test - verify CSV output
        df_result = pd.read_csv(target_csv_path)
        self.assertEqual(df_result.shape, (3, 3), "Should have 3 rows and 3 columns")
        self.assertEqual(list(df_result.columns), ['Name', 'Age', 'City'])
        self.assertEqual(df_result.iloc[0, 0], 'Alice')
        self.assertEqual(df_result.iloc[1, 1], 30)
        self.assertEqual(df_result.iloc[2, 2], 'Chicago')


    def test_mixed_format_operations(self):
        """Test 13: Mixed format operations (CSV + Excel) in single batch"""
        
        # Import required modules
        from openpyxl import Workbook
        import pandas as pd
        
        # Create CSV source
        csv_source_path = os.path.join(self.test_dir, "mixed_csv.csv")
        csv_data = {'Product': ['Laptop', 'Mouse'], 'Price': [1000, 25]}
        pd.DataFrame(csv_data).to_csv(csv_source_path, index=False)
        
        # Create Excel source
        excel_source_path = os.path.join(self.test_dir, "mixed_excel.xlsx")
        wb_excel = Workbook()
        ws_excel = wb_excel.active
        ws_excel.title = "Inventory"
        ws_excel['A1'] = "Category"
        ws_excel['A2'] = "Electronics"
        wb_excel.save(excel_source_path)
        
        # Create Excel target
        target_path = os.path.join(self.test_dir, "mixed_target.xlsx")
        wb_target = Workbook()
        ws_target = wb_target.active
        ws_target.title = "Combined"
        wb_target.save(target_path)
        
        # Test parameters - mixed format operations
        test_params = {
            "workspace": {
                "csv_src": csv_source_path,
                "excel_src": excel_source_path,
                "target": target_path
            },
            "operations": [
                {
                    "from": "[csv_src]A1:B3",  # CSV data
                    "to": "[target]Combined!A1:B3",
                    "save_as": "target",
                    "copy": ["values"]
                },
                {
                    "from": "[excel_src]Inventory!A1:A2",  # Excel data
                    "to": "[target]Combined!C1:C2",
                    "save_as": "target",
                    "copy": ["values"]
                }
            ]
        }
        
        # Execute the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test
        expected_result = """{
    "success": true,
    "operations_completed": 2,
    "results": [
        {
            "operation": 1,
            "success": true,
            "result": {
                "copied_cells": 6,
                "copy_types": [
                    "values"
                ]
            },
            "save_as": "target"
        },
        {
            "operation": 2,
            "success": true,
            "result": {
                "copied_cells": 2,
                "copy_types": [
                    "values"
                ]
            },
            "save_as": "target"
        }
    ],
    "files_saved": {
        "saved_files": [            
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\mixed_target.xlsx"
        ]
    }
}"""
        
        result_json = json.dumps(result, indent=4, sort_keys=True, ensure_ascii=False)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True, ensure_ascii=False)
        
        self.assertEqual(result_json, expected_json)
        
        # File Content Test
        target_content = self.read_excel_sheet_content(target_path, "Combined")
        expected_file_content = """SHEET: Combined
RANGE: A1:C3
DATA:
ROW1: ["Product", "Price", "Category"]
ROW2: ["Laptop", 1000, "Electronics"]
ROW3: ["Mouse", 25, NULL]"""
        
        self.assertEqual(target_content, expected_file_content)


    def test_complex_formulas(self):
        """Test 14: Complex formulas with functions like SUM, AVERAGE, VLOOKUP"""
        
        # Import required modules
        from openpyxl import Workbook
        
        # Create Excel target
        target_path = os.path.join(self.test_dir, "formulas_target.xlsx")
        wb_target = Workbook()
        ws_target = wb_target.active
        ws_target.title = "Calculations"
        wb_target.save(target_path)
        
        # Test parameters - complex formulas
        test_params = {
            "workspace": {
                "target": target_path
            },
            "operations": [
                {
                    "from": "=SUM(A1:A10)",
                    "to": "[target]Calculations!B1",
                    "save_as": "target",
                    "copy": ["formulas"]
                },
                {
                    "from": "=AVERAGE(B1:B5)",
                    "to": "[target]Calculations!B2",
                    "save_as": "target",
                    "copy": ["formulas"]
                },
                {
                    "from": "=IF(A1>0,\"Positive\",\"Negative\")",
                    "to": "[target]Calculations!C1",
                    "save_as": "target",
                    "copy": ["formulas"]
                },
                {
                    "from": "=CONCATENATE(\"Hello \",A1)",
                    "to": "[target]Calculations!D1",
                    "save_as": "target",
                    "copy": ["formulas"]
                }
            ]
        }
        
        # Execute the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test
        expected_result = """{
    "success": true,
    "operations_completed": 4,
    "results": [
        {
            "operation": 1,
            "success": true,
            "result": {
                "formula_set": 1,
                "formula": "=SUM(A1:A10)"
            },
            "save_as": "target"
        },
        {
            "operation": 2,
            "success": true,
            "result": {
                "formula_set": 1,
                "formula": "=AVERAGE(B1:B5)"
            },
            "save_as": "target"
        },
        {
            "operation": 3,
            "success": true,
            "result": {
                "formula_set": 1,
                "formula": "=IF(A1>0,\\"Positive\\",\\"Negative\\")"
            },
            "save_as": "target"
        },
        {
            "operation": 4,
            "success": true,
            "result": {
                "formula_set": 1,
                "formula": "=CONCATENATE(\\"Hello \\",A1)"
            },
            "save_as": "target"
        }
    ],
    "files_saved": {
        "saved_files": [
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\formulas_target.xlsx"
        ]
    }
}"""
        
        result_json = json.dumps(result, indent=4, sort_keys=True, ensure_ascii=False)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True, ensure_ascii=False)
        
        self.assertEqual(result_json, expected_json)
        
        # File Content Test
        target_content = self.read_excel_sheet_content(target_path, "Calculations")
        expected_file_content = """SHEET: Calculations
RANGE: A1:D2
DATA:
ROW1: [NULL, "=SUM(A1:A10)", "=IF(A1>0,"Positive","Negative")", "=CONCATENATE("Hello ",A1)"]
ROW2: [NULL, "=AVERAGE(B1:B5)", NULL, NULL]"""
        
        self.assertEqual(target_content, expected_file_content)


    def test_large_range_performance(self):
        """Test 15: Large range operations (A1:E100) for performance testing"""
        
        # Import required modules
        from openpyxl import Workbook
        
        # Create source with large data set
        source_path = os.path.join(self.test_dir, "large_source.xlsx")
        wb_source = Workbook()
        ws_source = wb_source.active
        ws_source.title = "LargeData"
        
        # Fill 100 rows x 5 columns with data
        for row in range(1, 101):
            for col in range(1, 6):
                ws_source.cell(row=row, column=col, value=f"R{row}C{col}")
        
        wb_source.save(source_path)
        
        # Create target file
        target_path = os.path.join(self.test_dir, "large_target.xlsx")
        wb_target = Workbook()
        ws_target = wb_target.active
        ws_target.title = "Results"
        wb_target.save(target_path)
        
        # Test parameters - large range
        test_params = {
            "workspace": {
                "source": source_path,
                "target": target_path
            },
            "operations": [
                {
                    "from": "[source]LargeData!A1:E100",
                    "to": "[target]Results!A1:E100",
                    "save_as": "target",
                    "copy": ["values"],
                    "insert": "replace"
                }
            ]
        }
        
        # Execute the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test
        expected_result = """{
    "success": true,
    "operations_completed": 1,
    "results": [
        {
            "operation": 1,
            "success": true,
            "result": {
                "copied_cells": 500,
                "copy_types": [
                    "values"
                ]
            },
            "save_as": "target"
        }
    ],
    "files_saved": {
        "saved_files": [
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\large_target.xlsx"
        ]
    }
}"""
        
        result_json = json.dumps(result, indent=4, sort_keys=True, ensure_ascii=False)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True, ensure_ascii=False)
        
        self.assertEqual(result_json, expected_json)
        
        # Performance validation - check that 500 cells were copied
        self.assertEqual(result["results"][0]["result"]["copied_cells"], 500)
        
        # Spot check some cells to ensure data was copied correctly
        import openpyxl
        wb_check = openpyxl.load_workbook(target_path)
        ws_check = wb_check["Results"]
        self.assertEqual(ws_check.cell(row=1, column=1).value, "R1C1")
        self.assertEqual(ws_check.cell(row=50, column=3).value, "R50C3")
        self.assertEqual(ws_check.cell(row=100, column=5).value, "R100C5")


    def test_special_characters_handling(self):
        """Test 16: Special characters in cell content (unicode, quotes, commas)"""
        
        # Import required modules
        from openpyxl import Workbook
        import pandas as pd
        
        # Create source CSV with special characters
        source_path = os.path.join(self.test_dir, "special_chars.csv")
        special_data = pd.DataFrame({
            "unicode": ["Привет мир 🌍", "こんにちは世界", "Hëllö Wörld"],
            "quotes": ['Text with "quotes"', "Text with 'apostrophe'", 'Mix "both" types'],
            "commas": ["Value,with,commas", "Another, separated, value", "Final,comma,test"]
        })
        special_data.to_csv(source_path, index=False, encoding='utf-8')
        
        # Create target Excel file
        target_path = os.path.join(self.test_dir, "special_target.xlsx")
        wb_target = Workbook()
        ws_target = wb_target.active
        ws_target.title = "SpecialChars"
        wb_target.save(target_path)
        
        # Test parameters
        test_params = {
            "workspace": {
                "source": source_path,
                "target": target_path
            },
            "operations": [
                {
                    "from": "[source]A1:C4",
                    "to": "[target]SpecialChars!A1:C4",
                    "save_as": "target",
                    "copy": ["values"],
                    "insert": "replace"
                }
            ]
        }
        
        # Execute the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test
        expected_result = """{
    "success": true,
    "operations_completed": 1,
    "results": [
        {
            "operation": 1,
            "success": true,
            "result": {
                "copied_cells": 12,
                "copy_types": [
                    "values"
                ]
            },
            "save_as": "target"
        }
    ],
    "files_saved": {
        "saved_files": [
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\special_target.xlsx"
        ]
    }
}"""
        
        result_json = json.dumps(result, indent=4, sort_keys=True, ensure_ascii=False)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True, ensure_ascii=False)
        
        self.assertEqual(result_json, expected_json)
        
        # Verify special characters are preserved
        import openpyxl
        wb_check = openpyxl.load_workbook(target_path)
        ws_check = wb_check["SpecialChars"]
        
        # Check unicode characters
        self.assertEqual(ws_check.cell(row=2, column=1).value, "Привет мир 🌍")
        # Check quotes handling
        self.assertEqual(ws_check.cell(row=2, column=2).value, 'Text with "quotes"')
        # Check commas handling
        self.assertEqual(ws_check.cell(row=2, column=3).value, "Value,with,commas")


    def test_insert_rows_and_columns_mode(self):
        """Test 17: Insert both rows and columns mode to test comprehensive insertion"""
        
        # Import required modules
        from openpyxl import Workbook
        
        # Create source file
        source_path = os.path.join(self.test_dir, "insert_source.xlsx")
        wb_source = Workbook()
        ws_source = wb_source.active
        ws_source.title = "SourceData"
        
        # Fill with new data to insert
        ws_source.cell(row=1, column=1, value="New1")
        ws_source.cell(row=1, column=2, value="New2")
        ws_source.cell(row=2, column=1, value="New3")
        ws_source.cell(row=2, column=2, value="New4")
        
        wb_source.save(source_path)
        
        # Create target file with existing data
        target_path = os.path.join(self.test_dir, "insert_target.xlsx")
        wb_target = Workbook()
        ws_target = wb_target.active
        ws_target.title = "TargetData"
        
        # Fill with existing data
        ws_target.cell(row=1, column=1, value="Existing1")
        ws_target.cell(row=1, column=2, value="Existing2")
        ws_target.cell(row=2, column=1, value="Existing3")
        ws_target.cell(row=2, column=2, value="Existing4")
        
        wb_target.save(target_path)
        
        # Test parameters - insert both rows and columns
        test_params = {
            "workspace": {
                "source": source_path,
                "target": target_path
            },
            "operations": [
                {
                    "from": "[source]SourceData!A1:B2",
                    "to": "[target]TargetData!A1:B2",
                    "save_as": "target",
                    "copy": ["values"],
                    "insert": ["rows", "columns"]
                }
            ]
        }
        
        # Execute the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test
        expected_result = """{
    "success": true,
    "operations_completed": 1,
    "results": [
        {
            "operation": 1,
            "success": true,
            "result": {
                "copied_cells": 4,
                "copy_types": [
                    "values"
                ],
                "insert_mode": [
                    "rows",
                    "columns"
                ]
            },
            "save_as": "target"
        }
    ],
    "files_saved": {
        "saved_files": [
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\insert_target.xlsx"
        ]
    }
}"""
        
        result_json = json.dumps(result, indent=4, sort_keys=True, ensure_ascii=False)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True, ensure_ascii=False)
        
        self.assertEqual(result_json, expected_json)
        
        # Verify that original data was shifted and new data inserted
        import openpyxl
        wb_check = openpyxl.load_workbook(target_path)
        ws_check = wb_check["TargetData"]
        
        # New data should be at A1:B2
        self.assertEqual(ws_check.cell(row=1, column=1).value, "New1")
        self.assertEqual(ws_check.cell(row=1, column=2).value, "New2")
        # Original data should be shifted to C3:D4
        self.assertEqual(ws_check.cell(row=3, column=3).value, "Existing1")
        self.assertEqual(ws_check.cell(row=4, column=4).value, "Existing4")


    def test_edge_cases(self):
        """Test 18: Edge cases - empty cells, None values, very long text"""
        
        # Import required modules
        from openpyxl import Workbook
        
        # Create source file with edge cases
        source_path = os.path.join(self.test_dir, "edge_source.xlsx")
        wb_source = Workbook()
        ws_source = wb_source.active
        ws_source.title = "EdgeData"
        
        # Fill with edge case data
        ws_source.cell(row=1, column=1, value="")  # Empty string
        ws_source.cell(row=1, column=2, value=None)  # None value
        ws_source.cell(row=1, column=3, value="A" * 1000)  # Very long text (1000 chars)
        ws_source.cell(row=2, column=1, value=0)  # Zero value
        ws_source.cell(row=2, column=2, value="   ")  # Whitespace only
        ws_source.cell(row=2, column=3, value="Multi\nLine\nText")  # Multi-line text
        
        wb_source.save(source_path)
        
        # Create target file
        target_path = os.path.join(self.test_dir, "edge_target.xlsx")
        wb_target = Workbook()
        ws_target = wb_target.active
        ws_target.title = "EdgeResults"
        wb_target.save(target_path)
        
        # Test parameters
        test_params = {
            "workspace": {
                "source": source_path,
                "target": target_path
            },
            "operations": [
                {
                    "from": "[source]EdgeData!A1:C2",
                    "to": "[target]EdgeResults!A1:C2",
                    "save_as": "target",
                    "copy": ["values"],
                    "insert": "replace"
                }
            ]
        }
        
        # Execute the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test
        expected_result = """{
    "success": true,
    "operations_completed": 1,
    "results": [
        {
            "operation": 1,
            "success": true,
            "result": {
                "copied_cells": 6,
                "copy_types": [
                    "values"
                ]
            },
            "save_as": "target"
        }
    ],
    "files_saved": {
        "saved_files": [
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\edge_target.xlsx"
        ]
    }
}"""
        
        result_json = json.dumps(result, indent=4, sort_keys=True, ensure_ascii=False)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True, ensure_ascii=False)
        
        self.assertEqual(result_json, expected_json)
        
        # Verify edge cases are handled correctly
        import openpyxl
        wb_check = openpyxl.load_workbook(target_path)
        ws_check = wb_check["EdgeResults"]
        
        # Check empty/None cells
        self.assertEqual(ws_check.cell(row=1, column=1).value, None)  # Empty becomes None
        self.assertEqual(ws_check.cell(row=1, column=2).value, None)  # None stays None
        # Check long text
        self.assertEqual(len(str(ws_check.cell(row=1, column=3).value)), 1000)
        # Check zero and whitespace
        self.assertEqual(ws_check.cell(row=2, column=1).value, 0)
        self.assertEqual(ws_check.cell(row=2, column=2).value, "   ")
        # Check multi-line text
        self.assertEqual(ws_check.cell(row=2, column=3).value, "Multi\nLine\nText")


    def test_workspace_alias_validation(self):
        """Test 19: Workspace alias validation and error handling"""
        
        # Test parameters with invalid workspace alias
        test_params = {
            "workspace": {
                "source": os.path.join(self.test_dir, "valid_source.xlsx")
            },
            "operations": [
                {
                    "from": "[nonexistent]Sheet1!A1:A2",  # Invalid alias
                    "to": "[source]Sheet1!B1:B2",
                    "save_as": "source",
                    "copy": ["values"],
                    "insert": "replace"
                }
            ]
        }
        
        # Execute the tool and expect failure
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test - should get error response
        expected_result = """{
    "success": false,
    "error": "Operation 1 failed: File ID 'nonexistent' not found in workspace",
    "operation": 1,
    "completed_operations": []
}"""
        
        result_json = json.dumps(result, indent=4, sort_keys=True, ensure_ascii=False)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True, ensure_ascii=False)
        
        self.assertEqual(result_json, expected_json)
        
        # Verify failure was handled gracefully
        self.assertFalse(result["success"])
        self.assertIn("File ID 'nonexistent' not found", result["error"])


    def test_comprehensive_mixed_scenario(self):
        """Test 20: Comprehensive mixed scenario - multiple files and formats"""
        
        # Import required modules
        from openpyxl import Workbook
        import pandas as pd
        
        # Create source Excel file
        excel_source_path = os.path.join(self.test_dir, "mixed_excel.xlsx")
        wb_excel = Workbook()
        ws1 = wb_excel.active
        ws1.title = "RawData"
        ws1.cell(row=1, column=1, value="Item")
        ws1.cell(row=1, column=2, value="Price")
        ws1.cell(row=2, column=1, value="Product A")
        ws1.cell(row=2, column=2, value=100)
        wb_excel.save(excel_source_path)
        
        # Create source CSV file
        csv_source_path = os.path.join(self.test_dir, "mixed_data.csv")
        csv_data = pd.DataFrame({
            "Category": ["Electronics", "Books"],
            "Count": [15, 25]
        })
        csv_data.to_csv(csv_source_path, index=False)
        
        # Create target Excel file
        target_path = os.path.join(self.test_dir, "mixed_final.xlsx")
        wb_target = Workbook()
        ws_summary = wb_target.active
        ws_summary.title = "Summary"
        wb_target.save(target_path)
        
        # Simple test parameters 
        test_params = {
            "workspace": {
                "excel_src": excel_source_path,
                "csv_src": csv_source_path,
                "target": target_path
            },
            "operations": [
                {
                    "from": "[excel_src]RawData!A1:B2",
                    "to": "[target]Summary!A1:B2",
                    "save_as": "target",
                    "copy": ["values"],
                    "insert": "replace"
                },
                {
                    "from": "[csv_src]A1:B3",
                    "to": "[target]Summary!D1:E3",
                    "save_as": "target",
                    "copy": ["values"],
                    "insert": "replace"
                }
            ]
        }
        
        # Execute the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test
        expected_result = """{
    "success": true,
    "operations_completed": 2,
    "results": [
        {
            "operation": 1,
            "success": true,
            "result": {
                "copied_cells": 4,
                "copy_types": [
                    "values"
                ]
            },
            "save_as": "target"
        },
        {
            "operation": 2,
            "success": true,
            "result": {
                "copied_cells": 6,
                "copy_types": [
                    "values"
                ]
            },
            "save_as": "target"
        }
    ],
    "files_saved": {
        "saved_files": [
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\mixed_final.xlsx"
        ]
    }
}"""
        
        result_json = json.dumps(result, indent=4, sort_keys=True, ensure_ascii=False)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True, ensure_ascii=False)
        
        self.assertEqual(result_json, expected_json)
        
        # Verify data was copied correctly
        import openpyxl
        wb_check = openpyxl.load_workbook(target_path)
        ws_summary = wb_check["Summary"]
        
        # Check Excel data at A1:B2
        self.assertEqual(ws_summary.cell(row=1, column=1).value, "Item")
        self.assertEqual(ws_summary.cell(row=2, column=2).value, 100)
        
        # Check CSV data at D1:E3  
        self.assertEqual(ws_summary.cell(row=1, column=4).value, "Category")
        self.assertEqual(ws_summary.cell(row=2, column=5).value, 15)


    def test_expressions_in_ranges(self):
        """Test expressions in ranges like A1:A{! env.VAR !} - from test_xls_fixes.py"""
        import tempfile
        
        # Set environment variable for test
        os.environ["ROW_COUNT"] = "5"
        
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create source CSV file
            source_csv = os.path.join(temp_dir, "source.csv")
            with open(source_csv, "w") as f:
                f.write("A,B,C\n1,2,3\n4,5,6\n7,8,9\n10,11,12\n13,14,15\n")
            
            # Create target Excel file
            target_xlsx = os.path.join(temp_dir, "target.xlsx")
            
            # Test copying with expression in range
            test_params = {
                "workspace": {
                    "source": source_csv,
                    "target": target_xlsx
                },
                "operations": [
                    {
                        "from": "[source]A1:C{! env.ROW_COUNT !}",  # Expression in range
                        "to": "[target]Sheet1!A1:C5",
                        "save_as": "target",
                        "copy": ["values"]
                    }
                ]
            }
            
            # Run the tool
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            try:
                result = loop.run_until_complete(main(test_params))
            finally:
                loop.close()
            
            # Verify success
            self.assertTrue(result["success"], f"Operation failed: {result.get('error', 'Unknown error')}")
            self.assertEqual(result["operations_completed"], 1)
            self.assertEqual(len(result["results"]), 1)
            self.assertEqual(result["results"][0]["result"]["copied_cells"], 15)  # 3x5 cells

    def test_insert_mode_in_response(self):
        """Test that insert_mode is included in response - from test_xls_fixes.py"""
        import tempfile
        
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create source CSV file
            source_csv = os.path.join(temp_dir, "source.csv")
            with open(source_csv, "w") as f:
                f.write("A,B\n1,2\n3,4\n")
            
            # Create target Excel file
            target_xlsx = os.path.join(temp_dir, "target.xlsx")
            
            # Test with insert mode different from "replace"
            test_params = {
                "workspace": {
                    "source": source_csv,
                    "target": target_xlsx
                },
                "operations": [
                    {
                        "from": "[source]A1:B3",
                        "to": "[target]Sheet1!A1:B3",
                        "save_as": "target",
                        "copy": ["values"],
                        "insert": ["rows"]  # Non-default insert mode
                    }
                ]
            }
            
            # Run the tool
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            try:
                result = loop.run_until_complete(main(test_params))
            finally:
                loop.close()
            
            # Verify success
            self.assertTrue(result["success"], f"Operation failed: {result.get('error', 'Unknown error')}")
            
            # Check if insert_mode is in the response - it's nested in result.result
            self.assertIn("results", result)
            self.assertGreater(len(result["results"]), 0)
            op_result = result["results"][0]
            self.assertIn("result", op_result)
            self.assertIn("insert_mode", op_result["result"])
            self.assertEqual(op_result["result"]["insert_mode"], ["rows"])

    def test_formatting_copy_no_styleproxy_error(self):
        """Test that formatting copy doesn't cause StyleProxy errors - from test_xls_fixes.py"""
        import tempfile
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        with tempfile.TemporaryDirectory() as temp_dir:
            # Create source Excel with formatting
            source_xlsx = os.path.join(temp_dir, "source.xlsx")
            target_xlsx = os.path.join(temp_dir, "target.xlsx")
            
            # Create a simple source workbook first
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            
            # Add data with formatting
            ws["A1"].value = "Header"
            ws["A1"].font = Font(bold=True)
            ws["A1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            ws["A2"].value = "Data"
            ws["A2"].font = Font(italic=True)
            
            wb.save(source_xlsx)
            wb.close()
            
            # Test copying with formatting
            test_params = {
                "workspace": {
                    "source": source_xlsx,
                    "target": target_xlsx
                },
                "operations": [
                    {
                        "from": "[source]Data!A1:A2",
                        "to": "[target]Sheet1!A1:A2",
                        "save_as": "target",
                        "copy": ["values", "formatting"]  # Include formatting
                    }
                ]
            }
            
            # Run the tool
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            try:
                result = loop.run_until_complete(main(test_params))
            finally:
                loop.close()
            
            # Verify success - no StyleProxy errors should occur
            self.assertTrue(result["success"], f"Operation failed: {result.get('error', 'Unknown error')}")
            self.assertEqual(result["operations_completed"], 1)
            self.assertEqual(len(result["results"]), 1)
            self.assertEqual(result["results"][0]["result"]["copied_cells"], 2)
            self.assertEqual(result["results"][0]["result"]["copy_types"], ["values", "formatting"])

    def test_single_cell_copy_fix(self):
        """Test copying single cell ranges (fix for 'Cell' object is not iterable)."""
        csv_file = os.path.join(self.test_dir, "single_cell_test.csv")
        xlsx_file = os.path.join(self.test_dir, "single_cell_target.xlsx")
        
        # Create test CSV with single cell data
        csv_content = "Name,Age,City\nJohn,25,NY\nJane,30,LA"
        with open(csv_file, 'w', encoding='utf-8') as f:
            f.write(csv_content)
        
        # Test copying single cell A1 to A1
        test_params = {
            "workspace": {
                "csv": csv_file,
                "xlsx": xlsx_file
            },
            "operations": [
                {
                    "from": "[csv]A1",
                    "to": "[xlsx]Sheet1!A1",
                    "save_as": "xlsx",
                    "copy": ["values"]
                }
            ]
        }
        
        # Run the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        self.assertTrue(result["success"])
        self.assertEqual(result["operations_completed"], 1)
        
        # Verify the single cell was copied
        wb = openpyxl.load_workbook(xlsx_file)
        ws = wb.active
        # Note: CSV might have "Name" or None in A1, both are valid test cases for single cell copy
        self.assertTrue(result["success"], "Single cell copy should succeed")
        self.assertIsNotNone(ws["A1"], "A1 should have a cell object")
        
        # Test copying single cell to different position with insert mode
        test_params2 = {
            "workspace": {
                "csv": csv_file,
                "xlsx": xlsx_file
            },
            "operations": [
                {
                    "from": "[csv]B2",  # "25"
                    "to": "[xlsx]Sheet1!C5",
                    "save_as": "xlsx",
                    "copy": ["values"],
                    "insert": ["rows"]
                }
            ]
        }
        
        # Run the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result2 = loop.run_until_complete(main(test_params2))
        finally:
            loop.close()
        
        self.assertTrue(result2["success"])
        self.assertIn("insert_mode", result2["results"][0]["result"])
        
        # Cleanup
        for f in [csv_file, xlsx_file]:
            if os.path.exists(f):
                os.unlink(f)


    def test_template_based_copying(self):
        """Test 25: Template-based copying - copy from source to template, save as different result file"""
        
        # Import required modules
        import pandas as pd
        from openpyxl import Workbook
        
        # Create CSV source file with data
        csv_source_path = os.path.join(self.test_dir, "template_data.csv")
        csv_data = {
            'Date': ['2025-03-17', '2025-03-18'], 
            'Users': [1, 5],
            'Engaged': [1, 3]
        }
        pd.DataFrame(csv_data).to_csv(csv_source_path, index=False)
        
        # Create Excel template file with existing structure
        template_path = os.path.join(self.test_dir, "report_template.xlsx")
        wb_template = Workbook()
        ws_template = wb_template.active
        ws_template.title = "Report"
        # Template has headers and some existing data
        ws_template['A1'] = "Monthly Report"
        ws_template['A3'] = "Date"
        ws_template['B3'] = "Active Users" 
        ws_template['C3'] = "Engaged Users"
        ws_template['A10'] = "Summary:"
        ws_template['B10'] = "=SUM(B4:B9)"
        wb_template.save(template_path)
        
        # Result file path (different from template)
        result_path = os.path.join(self.test_dir, "monthly_report.xlsx")
        
        # Test parameters - template-based copying
        test_params = {
            "workspace": {
                "csv_data": csv_source_path,
                "template": template_path,
                "result": result_path
            },
            "operations": [
                {
                    "from": "[csv_data]A2:C3",  # Skip CSV header row
                    "to": "[template]Report!A4:C5", 
                    "save_as": "result",
                    "copy": ["values"],
                    "insert": "replace"
                }
            ]
        }
        
        # Execute the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test
        expected_result = """{
    "success": true,
    "operations_completed": 1,
    "results": [
        {
            "operation": 1,
            "success": true,
            "result": {
                "copied_cells": 6,
                "copy_types": [
                    "values"
                ]
            },
            "save_as": "result"
        }
    ],
    "files_saved": {
        "saved_files": [
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\monthly_report.xlsx"
        ]
    }
}"""
        
        result_json = json.dumps(result, indent=4, sort_keys=True, ensure_ascii=False)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True, ensure_ascii=False)
        
        self.assertEqual(result_json, expected_json)
        
        # Verify template file was NOT modified
        self.assertTrue(os.path.exists(template_path), "Template file should still exist")
        template_content = self.read_excel_sheet_content(template_path, "Report")
        # Template should contain original data only
        self.assertIn("Monthly Report", template_content)
        self.assertIn("Summary:", template_content)
        self.assertNotIn("2025-03-17", template_content)  # Data should NOT be in template
        
        # Verify result file contains both template structure AND copied data
        self.assertTrue(os.path.exists(result_path), "Result file should be created")
        result_content = self.read_excel_sheet_content(result_path, "Report")
        # Result should have template structure + copied data
        self.assertIn("Monthly Report", result_content)  # From template
        self.assertIn("Summary:", result_content)        # From template
        self.assertIn("2025-03-17", result_content)      # From copied data
        self.assertIn("2025-03-18", result_content)      # From copied data


    def test_in_place_editing_same_file(self):
        """Test 26: In-place editing - when save_as points to same file as template"""
        
        # Import required modules
        import pandas as pd
        from openpyxl import Workbook
        
        # Create CSV source file
        csv_source_path = os.path.join(self.test_dir, "update_data.csv") 
        csv_data = {
            'Status': ['Updated', 'Modified'],
            'Value': [100, 200]
        }
        pd.DataFrame(csv_data).to_csv(csv_source_path, index=False)
        
        # Create Excel file that will be modified in-place
        excel_path = os.path.join(self.test_dir, "inplace_file.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        # Original data
        ws['A1'] = "Original Header"
        ws['A2'] = "Old Data"
        ws['B2'] = "Old Value"
        ws['A5'] = "Footer Info"
        wb.save(excel_path)
        
        # Test parameters - in-place editing (save_as = template file)
        test_params = {
            "workspace": {
                "csv_source": csv_source_path,
                "excel_file": excel_path
            },
            "operations": [
                {
                    "from": "[csv_source]A2:B3",  # Skip CSV header
                    "to": "[excel_file]Data!A2:B3",
                    "save_as": "excel_file",  # Same as template - in-place editing
                    "copy": ["values"],
                    "insert": "replace"
                }
            ]
        }
        
        # Execute the tool
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(main(test_params))
        finally:
            loop.close()
        
        # Approval Test
        expected_result = """{
    "success": true,
    "operations_completed": 1,
    "results": [
        {
            "operation": 1,
            "success": true,
            "result": {
                "copied_cells": 4,
                "copy_types": [
                    "values"
                ]
            },
            "save_as": "excel_file"
        }
    ],
    "files_saved": {
        "saved_files": [
            "C:\\\\Java\\\\CopipotTraining\\\\hello-langchain\\\\mcp_server\\\\tools\\\\lng_xls_batch\\\\stuff\\\\test_data\\\\inplace_file.xlsx"
        ]
    }
}"""
        
        result_json = json.dumps(result, indent=4, sort_keys=True, ensure_ascii=False)
        expected_json = json.dumps(json.loads(expected_result), indent=4, sort_keys=True, ensure_ascii=False)
        
        self.assertEqual(result_json, expected_json)
        
        # Verify the original file was modified in-place
        self.assertTrue(os.path.exists(excel_path), "Excel file should still exist")
        modified_content = self.read_excel_sheet_content(excel_path, "Data")
        
        # Should contain both original unchanged data and new copied data
        self.assertIn("Original Header", modified_content)  # Original header preserved
        self.assertIn("Footer Info", modified_content)      # Original footer preserved
        self.assertIn("Updated", modified_content)          # New data copied
        self.assertIn("Modified", modified_content)         # New data copied
        self.assertNotIn("Old Data", modified_content)      # Old data replaced


def run_single_test(test_name):
    """Run a single test by name."""
    suite = unittest.TestSuite()
    suite.addTest(TestLngXlsBatch(test_name))
    runner = unittest.TextTestRunner(verbosity=2)
    return runner.run(suite)


if __name__ == "__main__":
    # Run all tests
    unittest.main(verbosity=2)
